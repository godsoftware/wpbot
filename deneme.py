from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
import time
import pandas as pd
import os
import hashlib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
# ✅ Tarayıcı ayarları
options = Options()
options.add_argument("user-data-dir=C:\\Users\\ozkal.DESKTOP-8UHA164\\EdgeBotProfile")
options.add_argument("profile-directory=Profile 1")
options.add_argument("--disable-extensions")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Edge(options=options)
driver.get("https://web.whatsapp.com")
print("🔓 Lütfen sohbetlerin yüklenmesini bekleyin...")
time.sleep(10)

# ✅ Excel dosyası varsa önceki kayıtları oku
excel_path = "whatsapp_mesajlar.xlsx"
previous_hashes = set()
if os.path.exists(excel_path):
    old_df = pd.read_excel(excel_path)
    for _, row in old_df.iterrows():
        key = f"{row['Numara']}_{row['Tarih']}_{row['Mesaj']}"
        hash_val = hashlib.md5(key.encode()).hexdigest()
        previous_hashes.add(hash_val)
else:
    old_df = pd.DataFrame()

# 📋 Yeni mesajlar burada toplanacak
all_messages = []

chat_boxes = driver.find_elements(By.XPATH, '//div[@aria-label="Sohbet listesi"]//div[contains(@style, "translateY")]')

for index, chat in enumerate(chat_boxes):
    try:
        chat_list = driver.find_elements(By.XPATH, '//div[@aria-label="Sohbet listesi"]//div[contains(@style, "translateY")]')
        chat = chat_list[index]

        number = chat.find_element(By.XPATH, './/span[contains(text(), "+90")]').text.strip()
        date = chat.find_element(By.XPATH, './/div[contains(@class, "_ak8i")]').text.strip()
        print(f"\n📨 [{index+1}] Sohbet: {number} | Tarih: {date}")

        chat.click()
        time.sleep(3)

        # Scroll alanı
        scroll_panel = driver.find_element(By.XPATH, '//div[contains(@class, "_akbu")]')
        last_height = driver.execute_script("return arguments[0].scrollHeight", scroll_panel)
        same_count = 0

        print("🔁 Mesaj geçmişi yükleniyor...")
        while True:
            driver.execute_script("arguments[0].scrollTop = 0", scroll_panel)
            time.sleep(2)
            new_height = driver.execute_script("return arguments[0].scrollHeight", scroll_panel)
            if new_height == last_height:
                same_count += 1
                if same_count >= 2:
                    break
            else:
                same_count = 0
                last_height = new_height

        print("✅ Tüm mesajlar yüklendi.")

        # Mesajları al
        message_blocks = driver.find_elements(By.XPATH, '//div[@class="copyable-text"]')
        for msg in message_blocks:
            try:
                timestamp = msg.get_attribute("data-pre-plain-text").strip()
                text_elems = msg.find_elements(By.XPATH, './/span[contains(@class, "selectable-text")]')
                for span in text_elems:
                    text = span.text.strip()
                    if text:
                        # Göndereni ayır
                        sender = "Bilinmiyor"
                        if "]" in timestamp:
                            sender_info = timestamp.split("]")[-1].strip().replace(":", "")
                            if sender_info == number:
                                sender = "GELEN"
                            else:
                                sender = "GİDEN"
                        time_only = timestamp.split("]")[0].replace("[", "")

                        # ✅ Sadece GELEN mesajlar işlensin
                        if sender == "GELEN":
                            key = f"{number}_{time_only}_{text}"
                            hash_val = hashlib.md5(key.encode()).hexdigest()
                            if hash_val not in previous_hashes:
                                all_messages.append({
                                    "Sohbet No": index + 1,
                                    "Numara": number,
                                    "Tarih": time_only,
                                    "Gönderen": sender,
                                    "Mesaj": text
                                })
                                previous_hashes.add(hash_val)
            except:
                continue

        time.sleep(1)

    except Exception as e:
        print(f"⚠️ {index+1}. sohbet atlandı: {e}")
        continue

if all_messages:
    df_new = pd.DataFrame(all_messages)

    if not old_df.empty:
        number_order = old_df['Numara'].drop_duplicates().tolist()
        new_numbers = df_new['Numara'].drop_duplicates().tolist()
        for num in new_numbers:
            if num not in number_order:
                number_order.append(num)

        df_combined = pd.concat([old_df, df_new], ignore_index=True)

        df_final = df_combined.copy()
        df_final["Tarih"] = pd.to_datetime(df_final["Tarih"], format="%H:%M, %d.%m.%Y", errors="coerce")
        df_final['Numara'] = pd.Categorical(df_final['Numara'], categories=number_order, ordered=True)
        df_final = df_final.sort_values(['Numara', 'Tarih'])
    else:
        df_final = df_new.sort_values(by="Tarih")

    # ✅ Sohbet No eşlemesini KORUYARAK her numarayı BLOK halinde sırala
    sohbet_map = {}
    if not old_df.empty and 'Sohbet No' in old_df.columns:
        sohbet_map = dict(zip(old_df['Numara'], old_df['Sohbet No']))

    max_sohbet_no = max(sohbet_map.values()) if sohbet_map else 0
    ordered_numbers = df_final['Numara'].drop_duplicates().tolist()
    for num in ordered_numbers:
        if num not in sohbet_map:
            max_sohbet_no += 1
            sohbet_map[num] = max_sohbet_no

    df_final['Sohbet No'] = df_final['Numara'].map(sohbet_map)

    # 📌 Numara sıralamasını koruyarak tekrar sırala
    df_final['Numara'] = pd.Categorical(df_final['Numara'], categories=ordered_numbers, ordered=True)
    df_final = df_final.sort_values(['Numara', 'Tarih'])



    # 🕒 Şu anki zaman
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 🧼 Numara sütunu string olsun (emin olmak için)
    df_final["Numara"] = df_final["Numara"].astype(str)

    # 🔁 "Son Kontrol" sütunu oluştur ve tümünü boş bırak
    df_final["Son Kontrol"] = ""

    # ✅ Her numaranın son satırına şu anki zamanı yaz
    for num in df_final["Numara"].unique():
        numara_rows = df_final[df_final["Numara"] == num]
        if not numara_rows.empty:
            last_index = numara_rows.index[-1]
            df_final.at[last_index, "Son Kontrol"] = now

    # 💾 Excel'e yaz
    df_final.to_excel(excel_path, index=False)
    print(f"\n📁 Yeni {len(all_messages)} gelen mesaj eklendi. Dosya güncellendi.")



    # Renk listesi (farklı numaralar için sırayla uygulanır)
    colors = ['FFFFCC', 'CCFFFF', 'FFCCCC', 'E0FFFF', 'FFE4E1', 'F5F5DC', 'E6E6FA', 'F0FFF0', 'FFF0F5']

    wb = load_workbook(excel_path)
    ws = wb.active

    # Numara sütununun index’ini bul
    numara_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Numara":
            numara_col = idx
            break

    if numara_col:
        current_color_index = 0
        previous_numara = None

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
            numara = row[numara_col - 1].value
            if numara != previous_numara:
                current_color_index = (current_color_index + 1) % len(colors)
                previous_numara = numara

            fill = PatternFill(start_color=colors[current_color_index],
                               end_color=colors[current_color_index],
                               fill_type="solid")

            for cell in row:
                cell.fill = fill

    wb.save(excel_path)
    print("🎨 Excel dosyasında her numara bloğu farklı renge boyandı.")

else:
    print("📁 Yeni gelen mesaj bulunamadı. Dosya zaten güncel.")



driver.quit()


